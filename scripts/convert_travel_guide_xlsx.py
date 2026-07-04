from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - friendly CLI failure
    raise SystemExit("缺少 openpyxl，无法读取 xlsx。请先安装 openpyxl 或使用 Codex bundled Python。") from exc


CATEGORY_COLUMNS = {
    "transport": "交通安排",
    "attractions": "必打卡景点",
    "food": "美食推荐",
    "tips": "实用小贴士",
}

CATEGORY_TITLES = {
    "transport": "交通安排与到达方式",
    "attractions": "必打卡景点推荐",
    "food": "美食体验推荐",
    "tips": "实用旅行提醒",
    "itinerary": "旅行计划素材汇总",
}

MIN_CONTENT_LENGTH = 80

STOP_WORDS = {
    "中国",
    "省",
    "市",
    "县",
    "区",
    "自治州",
    "风景区",
    "景区",
    "古城",
    "古镇",
    "旅游",
    "景点",
}

PROVINCE_NAMES = [
    "北京",
    "天津",
    "上海",
    "重庆",
    "河北",
    "山西",
    "辽宁",
    "吉林",
    "黑龙江",
    "江苏",
    "浙江",
    "安徽",
    "福建",
    "江西",
    "山东",
    "河南",
    "湖北",
    "湖南",
    "广东",
    "海南",
    "四川",
    "贵州",
    "云南",
    "陕西",
    "甘肃",
    "青海",
    "台湾",
    "内蒙古",
    "广西",
    "西藏",
    "宁夏",
    "新疆",
    "香港",
    "澳门",
]

PROVINCE_ALIASES = {
    "广西壮族自治区": "广西",
    "内蒙古自治区": "内蒙古",
    "西藏自治区": "西藏",
    "宁夏回族自治区": "宁夏",
    "新疆维吾尔自治区": "新疆",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_destination(value: str) -> str:
    text = clean_text(value)
    text = text.replace("旅游景点", "").replace("旅游区", "")
    text = re.sub(r"\s+", "", text)
    return text


def slugify(text: str) -> str:
    mapping = {
        "北京": "beijing",
        "上海": "shanghai",
        "重庆": "chongqing",
        "成都": "chengdu",
        "杭州": "hangzhou",
        "桂林": "guilin",
        "阳朔": "yangshuo",
        "丽江": "lijiang",
        "青岛": "qingdao",
        "张家界": "zhangjiajie",
        "凤凰": "fenghuang",
        "婺源": "wuyuan",
        "千岛湖": "qiandaohu",
        "拉萨": "lhasa",
        "敦煌": "dunhuang",
        "厦门": "xiamen",
        "大理": "dali",
        "南京": "nanjing",
        "苏州": "suzhou",
        "西安": "xian",
        "武汉": "wuhan",
        "长沙": "changsha",
        "昆明": "kunming",
        "三亚": "sanya",
    }
    for zh, slug in mapping.items():
        if zh in text:
            return slug
    ascii_slug = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    if ascii_slug:
        return ascii_slug[:40]
    return "dest_" + str(abs(hash(text)))[:8]


def detect_province(destination: str) -> str:
    for full, short in PROVINCE_ALIASES.items():
        if full in destination:
            return short
    for name in PROVINCE_NAMES:
        if name in destination:
            return name
    return ""


def extract_tags(destination: str, text: str, limit: int = 8) -> list[str]:
    candidates: list[str] = []
    for token in re.split(r"[，,。、；;：:\s/]+", destination):
        token = token.strip("《》“”\"'（）()[]【】")
        if 2 <= len(token) <= 12 and token not in STOP_WORDS:
            candidates.append(token)

    patterns = [
        r"[\u4e00-\u9fffA-Za-z0-9·]{2,16}(?:古城|古镇|雪山|湖|岛|寺|塔|山|河|江|街|巷|桥|瀑布|草原|湿地|公园|博物馆|景区|风景区|森林公园)",
        r"[\u4e00-\u9fffA-Za-z0-9·]{2,12}(?:火锅|米粉|烤鱼|鱼头汤|小面|酸汤鱼|腊排骨|烧饼|土鸡煲)",
    ]
    for pattern in patterns:
        candidates.extend(re.findall(pattern, text))

    seen = set()
    tags = []
    for item in candidates:
        cleaned = item.strip("《》“”\"'（）()[]【】")
        if cleaned and cleaned not in seen and cleaned not in STOP_WORDS:
            tags.append(cleaned)
            seen.add(cleaned)
        if len(tags) >= limit:
            break
    return tags or [destination[:12]]


def merge_texts(values: list[str], max_chars: int = 1300) -> str:
    seen = set()
    chunks = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        chunks.append(text)
        seen.add(text)
        if sum(len(chunk) for chunk in chunks) >= max_chars:
            break
    merged = "\n".join(chunks)
    return merged[:max_chars].strip()


def make_item(
    destination: str,
    province: str,
    category: str,
    content: str,
    index: int,
    source_rows: list[int],
    updated_at: str,
) -> dict[str, Any]:
    tags = extract_tags(destination, content)
    title = f"{destination}{CATEGORY_TITLES[category]}"
    return {
        "id": f"{slugify(destination)}_{category}_excel_{index:03d}",
        "city": destination,
        "province": province,
        "country": "中国",
        "category": category,
        "title": title,
        "tags": tags,
        "suitable_for": ["自由行游客", "第一次到访", "行程规划参考"],
        "duration": "行前参考",
        "best_season": "",
        "source_type": "excel_import_draft",
        "source_url": "",
        "license": "user_provided_excel",
        "updated_at": updated_at,
        "review_status": "pending",
        "source_file": "travel_guide.xlsx",
        "source_rows": source_rows,
        "content": content,
    }


def make_itinerary_content(destination: str, grouped: dict[str, list[str]]) -> str:
    parts = []
    attractions = merge_texts(grouped.get("必打卡景点", []), max_chars=520)
    food = merge_texts(grouped.get("美食推荐", []), max_chars=360)
    transport = merge_texts(grouped.get("交通安排", []), max_chars=300)
    tips = merge_texts(grouped.get("实用小贴士", []), max_chars=300)
    if attractions:
        parts.append(f"景点素材：{attractions}")
    if food:
        parts.append(f"美食素材：{food}")
    if transport:
        parts.append(f"交通素材：{transport}")
    if tips:
        parts.append(f"提醒素材：{tips}")
    if not parts:
        return ""
    return f"{destination}的行程规划可以综合以下素材生成，正式入库前建议人工审核并拆分为明确天数路线。\n" + "\n".join(parts)


def read_workbook(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook.active
    headers = [clean_text(cell.value) for cell in sheet[1]]
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = dict(zip(headers, row))
        if not any(clean_text(value) for value in values.values()):
            continue
        values["_row_number"] = row_number
        rows.append(values)
    return headers, rows


def convert(input_path: Path, output_path: Path, limit_destinations: int | None = None) -> list[dict[str, Any]]:
    headers, rows = read_workbook(input_path)
    required = {"目的地", *CATEGORY_COLUMNS.values()}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"Excel 缺少列：{', '.join(missing)}")

    by_destination: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        destination = normalize_destination(clean_text(row.get("目的地")))
        if not destination:
            continue
        by_destination[destination].append(row)

    updated_at = date.today().isoformat()
    items: list[dict[str, Any]] = []
    for index, (destination, destination_rows) in enumerate(by_destination.items(), start=1):
        if limit_destinations and index > limit_destinations:
            break
        province = detect_province(destination)
        source_rows = [int(row["_row_number"]) for row in destination_rows]

        grouped: dict[str, list[str]] = defaultdict(list)
        for row in destination_rows:
            for column in CATEGORY_COLUMNS.values():
                text = clean_text(row.get(column))
                if text:
                    grouped[column].append(text)

        for category, column in CATEGORY_COLUMNS.items():
            content = merge_texts(grouped.get(column, []))
            if len(content) >= MIN_CONTENT_LENGTH:
                items.append(make_item(destination, province, category, content, index, source_rows, updated_at))

        itinerary_content = make_itinerary_content(destination, grouped)
        if len(itinerary_content) >= MIN_CONTENT_LENGTH:
            items.append(make_item(destination, province, "itinerary", itinerary_content, index, source_rows, updated_at))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    return items


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert travel_guide.xlsx to knowledge draft JSON.")
    parser.add_argument("--input", default="travel_guide.xlsx", help="Excel 文件路径")
    parser.add_argument("--output", default="data/knowledge_drafts/travel_guide_import.json", help="输出草稿 JSON 路径")
    parser.add_argument("--limit-destinations", type=int, default=None, help="仅转换前 N 个目的地，便于试跑")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    items = convert(input_path, output_path, args.limit_destinations)
    destinations = len({item["city"] for item in items})
    print(f"转换完成：{input_path} -> {output_path}")
    print(f"目的地数量：{destinations}")
    print(f"知识条目数量：{len(items)}")
    print("状态：review_status=pending，尚未合并正式知识库。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
